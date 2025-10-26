import os
import random
import uuid

import pendulum
import snoop
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

import config
import locator
from db import pool

os.makedirs("logs", exist_ok=True)
now = pendulum.now()
format_name = now.format("YYYY-MM-DD_HH-mm-ss")
snoop.install(out=f"logs/{format_name}.log")


@snoop
def random_date(start_date, end_date):
    days_diff = (end_date - start_date).days
    random_days = random.randint(0, days_diff)
    random_date = start_date.add(days=random_days)
    formatted_date = random_date.format("DD/MM/YYYY")

    return formatted_date


@snoop
def run():
    with sync_playwright() as p:
        # Config browser and context page
        browser = p.chromium.launch(headless=config.HEADLESS, slow_mo=config.SLOW_MO)
        context = browser.new_context(
            record_video_dir=f"screen-record/{format_name}",
            record_video_size={
                "width": config.VIDEO_WIDTH_SIZE,
                "height": config.VIDEO_HEIGHT_SIZE,
            },
        )

        # Trace activity
        context.tracing.start(screenshots=True, snapshots=True, sources=True)

        page = context.new_page()
        page.set_viewport_size(
            {
                "width": config.VIEW_PORT_WIDTH_SIZE,
                "height": config.VIEW_PORT_HEIGHT_SIZE,
            }
        )
        page.goto(config.URL)

        # Homepage
        page.click(locator.MENU_TOGGLE)
        page.screenshot(
            path=f"screenshot/{format_name}/001-homepage.png", full_page=True
        )
        page.click(locator.LOGIN_MENU_TEXT)

        # Login page
        username = page.input_value(locator.USERNAME_TEXT)
        password = page.input_value(locator.PASSWORD_TEXT)
        page.fill(locator.USERNAME_INPUT_TEXT, username)
        page.fill(locator.PASSWORD_INPUT_TEXT, password)
        page.screenshot(
            path=f"screenshot/{format_name}/002-loginpage.png", full_page=True
        )
        page.click(locator.LOGIN_BTN)

        # Appointment page
        page.select_option(
            locator.FACILITY_DROPDOWNLIST, value="Seoul CURA Healthcare Center"
        )
        page.click(locator.APPLY_FOR_HOSTPITAL_READMISSION_CHECKBOX)
        page.click(locator.HEALTHCARE_PROGRAM_MEDICAID_RB)
        page.type(
            locator.VISIT_DATE_CALENDAR,
            random_date(pendulum.date(2020, 1, 1), pendulum.date(2025, 12, 31)),
        )
        page.keyboard.press("Tab")
        page.fill(
            locator.COMMENT_INPUT_TEXT,
            random.choice(
                ["Apping Ganteng 😉", "Alek Sayang Ibu ❤️", "Tangan Yoga bau rokok 🚬"]
            ),
        )
        page.screenshot(
            path=f"screenshot/{format_name}/003-appointment-form.png", full_page=True
        )
        page.click(locator.BOOK_APPOINTMENT_BTN)
        page.screenshot(
            path=f"screenshot/{format_name}/004-appointment-confirmation.png",
            full_page=True,
        )

        # Save appointment to database
        try:
            with pool.connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO appointments(id, facility, apply_for_hospital_readmission, healthcare_program, visit_date, comment, status)
                        VALUES(%s, %s, %s, %s, %s, %s, %s);
                        """,
                        (
                            uuid.uuid4(),
                            page.inner_text(locator.FACILITY_TEXT),
                            page.inner_text(
                                locator.APPLY_FOR_HOSTPITAL_READMISSION_TEXT
                            ),
                            page.inner_text(locator.HEALTHCARE_PROGRAM_TEXT),
                            page.inner_text(locator.VISIT_DATE_TEXT),
                            page.inner_text(locator.COMMENT_TEXT),
                            "Appointment Success",
                        ),
                    )
                    cur.close()
                conn.commit()
        except Exception as e:
            print(f"Caught: {e}")
            raise

        # Save appointment to excel file
        if not os.path.exists("appointment.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.title = "appointment"
            ws["A1"] = "Facility"
            ws["B1"] = "Apply for hospital readmission"
            ws["C1"] = "Healthcare Program"
            ws["D1"] = "Visit Date"
            ws["E1"] = "Comment"
            ws["F1"] = "Status"
            wb.save("appointment.xlsx")
        else:
            wb = load_workbook("appointment.xlsx")
            ws = wb["appointment"]

        last_row = ws.max_row + 1
        ws["A" + str(last_row)] = page.inner_text(locator.FACILITY_TEXT)
        ws["B" + str(last_row)] = page.inner_text(
            locator.APPLY_FOR_HOSTPITAL_READMISSION_TEXT
        )
        ws["C" + str(last_row)] = page.inner_text(locator.HEALTHCARE_PROGRAM_TEXT)
        ws["D" + str(last_row)] = page.inner_text(locator.VISIT_DATE_TEXT)
        ws["E" + str(last_row)] = page.inner_text(locator.COMMENT_TEXT)
        ws["F" + str(last_row)] = "Appointment Success"
        wb.save("appointment.xlsx")
        wb.close()

        page.click(locator.GOTO_HOMEPAGE_BTN)
        page.click(locator.MENU_TOGGLE)
        page.click(locator.LOGOUT_MENU_TEXT)

        context.tracing.stop(path="trace.zip")

        context.close()
        browser.close()


if __name__ == "__main__":
    run()
